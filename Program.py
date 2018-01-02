import os
import xlwt
import datetime, xlrd
import time
import pandas as pd 
import operator

from classes import Paddler, Boat
from openpyxl import load_workbook

# cwd = os.getcwd()
# print(cwd)



# print(os.listdir('.'))

# Assign filename to 'file'
# file = 'List.xlsx'

# Load Spreadsheet
# xl = pd.ExcelFile(file)

# Print the sheet names
# print(xl.sheet_names)

# Load a sheet into a DataFrame by name: df1
# df1 = xl.parse('Sheet1')

def TotalWeight(Side):
	weight = 0;
	for i in range(0,len(Side)):
		weight = weight + Side[i].Weight
	return weight

wb = load_workbook('./List.xlsx')

# print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet1')

sheet.title

# anotherSheet = wb.active

# anotherSheet

numPaddlers = input('Input number of Paddlers: ')

# sheet['A1'].value

# All Paddler Information passed into AllPaddlers
AllPaddlers = []

# TimeTrialFormat = wb.add_format({'Time Format': 'Min:Sec.MSec'})

for i in range (2,numPaddlers + 2):
	name = sheet.cell(row=i, column=1).value
	weight = sheet.cell(row=i, column=2).value
	gender = sheet.cell(row=i, column=3).value
	side = sheet.cell(row=i, column=4).value
	TimeInput = str(sheet.cell(row=i, column=5).value)
	
	Hours, Minutes, Secs = TimeInput.split(":", 2)
	# print(Minutes, Secs)
	time = float(Minutes) * 60 + float(Secs)
	# T = time.strptime(str(sheet.cell(row=i, column=5).value), '%H:%M:%S.%f')
	# time = datetime.timedelta(sheet.cell(row=i, column=5).value)

	# time = sheet.cell(row=i, column=5).value
	notes = sheet.cell(row=i, column=6).value
	newPaddler = Paddler(name, weight, gender, side, time, notes)
	AllPaddlers.append(newPaddler)



Boats = []
numBoats = input('Input number of Boats: ')

B = 0
while B < numBoats:
	# Sorts Paddlers by Trial Times
	AllPaddlers.sort(key=operator.attrgetter('TimeTrial'))
	LeftPaddlers = []
	RightPaddlers = []
	Both = []

	# Seperates into Left, Right and Both Side preferences
	for i in range (0, len(AllPaddlers)):
		if AllPaddlers[i].Side == 'L':
			LeftPaddlers.append(AllPaddlers[i])
		elif AllPaddlers[i].Side == 'R':
			RightPaddlers.append(AllPaddlers[i])
		elif AllPaddlers[i].Side == 'B':
			Both.append(AllPaddlers[i])

	print('Num Paddlers: ', len(AllPaddlers))

	while 1:
		SplitRatio = input('Choose option for Gender Split Ratio (Male:Female) for boat: \n1. 10:10 \n2. 12:6 ')
		
		if SplitRatio == 1:
			numMales = 10
			numFemales = 10
			BoatSize = 20
			break
		elif SplitRatio == 2:
			numMales = 12
			numFemales = 6
			BoatSize = 18
			break
		else:
			print('Invalid Selection, Please try again.')


	BoatLeft = []
	BoatRight = []
	BoatBoth = []

	MaleCounter = 0
	FemaleCounter = 0
	i = 0
	while i < len(AllPaddlers):
		if len(BoatRight) + len(BoatLeft) + len(BoatBoth) < numMales + numFemales:
			if AllPaddlers[i].Gender == 'M' and MaleCounter < numMales:
				# print(AllPaddlers[i].printPaddler())
				print('Male ', MaleCounter + 1, ' Added')
				if AllPaddlers[i].Side == 'R' and len(BoatRight) < BoatSize / 2:
					BoatRight.append(AllPaddlers[i])
					AllPaddlers.remove(AllPaddlers[i])
					MaleCounter += 1
					i = 0
					continue
				elif AllPaddlers[i].Side == 'L' and len(BoatLeft) < BoatSize / 2:
					BoatLeft.append(AllPaddlers[i])
					AllPaddlers.remove(AllPaddlers[i])
					MaleCounter += 1
					i = 0
					continue
				elif AllPaddlers[i].Side == 'B' and len(BoatBoth) < BoatSize:
					BoatBoth.append(AllPaddlers[i])
					AllPaddlers.remove(AllPaddlers[i])
					MaleCounter += 1
					i = 0
					continue
				print('Male ', MaleCounter, ' Added')
				# i += 1
				# print(MaleCounter)
			elif AllPaddlers[i].Gender == 'F' and FemaleCounter < numFemales:
				# print(AllPaddlers[i].printPaddler())
				print('Female ', FemaleCounter + 1, ' Added')
				if AllPaddlers[i].Side == 'R' and len(BoatRight) < BoatSize / 2:
					BoatRight.append(AllPaddlers[i])
					AllPaddlers.remove(AllPaddlers[i])
					FemaleCounter += 1
					i = 0
					continue
				elif AllPaddlers[i].Side == 'L' and len(BoatLeft) < BoatSize / 2:
					BoatLeft.append(AllPaddlers[i])
					AllPaddlers.remove(AllPaddlers[i])
					FemaleCounter += 1
					i = 0
					continue
				elif AllPaddlers[i].Side == 'B' and len(BoatBoth) < BoatSize:
					BoatBoth.append(AllPaddlers[i])
					AllPaddlers.remove(AllPaddlers[i])
					FemaleCounter += 1
					i = 0 
					continue
				# i += 1
				print('Female ', FemaleCounter, ' Added')
		i += 1  

	i = 0		
	while len(BoatBoth) != 0:
		if len(BoatRight) < BoatSize / 2:
			BoatRight.append(BoatBoth[i])
			BoatBoth.remove(BoatBoth[i])
		elif len(BoatLeft) < BoatSize / 2:
			BoatLeft.append(BoatBoth[i])
			BoatBoth.remove(BoatBoth[i])

	Substitutes = []
	AllPaddlers.reverse()
	while len(Substitutes) < 4 and len(AllPaddlers) > 0:
		Substitutes.append(AllPaddlers.pop())
	AllPaddlers.reverse()

	BoatLeft.sort(key=operator.attrgetter('Weight'))
	BoatRight.sort(key=operator.attrgetter('Weight'))

	WeightSortedLeft = []
	WeightSortedRight = []

	for i in range(0, len(BoatLeft)):
		if i % 2 == 0:
			WeightSortedLeft.append(BoatLeft[len(BoatLeft) - i - 1])
		else:
			WeightSortedLeft.insert(0, BoatLeft[len(BoatLeft) - i - 1])


	for i in range(0, len(BoatRight)):
		if i % 2 == 0:
			WeightSortedRight.append(BoatRight[len(BoatRight) - i - 1])
		else:
			WeightSortedRight.insert(0, BoatRight[len(BoatRight) - i - 1])


	Boats.append(Boat(B + 1, WeightSortedLeft, WeightSortedRight, Substitutes))
	# print('Left Side Total Weight:', TotalWeight(Boats[B].LeftSide))
	B += 1



for i in range(0, len(Boats)):
	Boats[i].printBoat()


book = xlwt.Workbook(encoding='utf-8')

sheet1 = book.add_sheet("Python Sheet 1")

# Row/Column
RowLastUsed = 0
for i in range(0, len(Boats)):	# i = Number of Boats
	BoatNumber = 'Boat ' + str(i + 1) 
	sheet1.write(RowLastUsed, 0, BoatNumber)
	RowLastUsed += 1
	sheet1.write(RowLastUsed, 0, 'Left Side Paddlers: ')
	sheet1.write(RowLastUsed, 7, 'Right Side Paddlers: ')
	RowLastUsed += 1
	sheet1.write(RowLastUsed, 0, 'Name: ')
	sheet1.write(RowLastUsed, 1, 'Weight: ')
	sheet1.write(RowLastUsed, 2, 'Gender: ')
	sheet1.write(RowLastUsed, 3, 'Side: ')
	sheet1.write(RowLastUsed, 4, 'Trial Time: ')
	sheet1.write(RowLastUsed, 5, 'Notes: ')
	sheet1.write(RowLastUsed, 7, 'Name: ')
	sheet1.write(RowLastUsed, 8, 'Weight: ')
	sheet1.write(RowLastUsed, 9, 'Gender: ')
	sheet1.write(RowLastUsed, 10, 'Side: ')
	sheet1.write(RowLastUsed, 11, 'Trial Time: ')
	sheet1.write(RowLastUsed, 12, 'Notes: ')
	RowLastUsed += 1
	for j in range(0, len(Boats[i].LeftSide)):	# j = Sides titles
		sheet1.write(j + RowLastUsed, 0, Boats[i].LeftSide[j].Name)
		sheet1.write(j + RowLastUsed, 1, Boats[i].LeftSide[j].Weight)
		sheet1.write(j + RowLastUsed, 2, Boats[i].LeftSide[j].Gender)
		sheet1.write(j + RowLastUsed, 3, Boats[i].LeftSide[j].Side)
		sheet1.write(j + RowLastUsed, 4, Boats[i].LeftSide[j].TimeTrial)
		sheet1.write(j + RowLastUsed, 5, Boats[i].LeftSide[j].Notes)
	for j in range(0, len(Boats[i].RightSide)):
		sheet1.write(j + RowLastUsed, 7, Boats[i].RightSide[j].Name)
		sheet1.write(j + RowLastUsed, 8, Boats[i].RightSide[j].Weight)
		sheet1.write(j + RowLastUsed, 9, Boats[i].RightSide[j].Gender)
		sheet1.write(j + RowLastUsed, 10, Boats[i].RightSide[j].Side)
		sheet1.write(j + RowLastUsed, 11, Boats[i].RightSide[j].TimeTrial)
		sheet1.write(j + RowLastUsed, 12, Boats[i].RightSide[j].Notes)
	
	if len(Boats[i].LeftSide) > len(Boats[i].RightSide):
		RowLastUsed += len(Boats[i].LeftSide)
	else:
		RowLastUsed += len(Boats[i].RightSide)

	sheet1.write(RowLastUsed, 0, 'Total Left Weight: ')
	LeftTotalWeight = TotalWeight(Boats[i].LeftSide)
	sheet1.write(RowLastUsed, 1, LeftTotalWeight)
	sheet1.write(RowLastUsed, 7, 'Total Right Weight: ')
	RightTotalWeight = TotalWeight(Boats[i].RightSide)
	sheet1.write(RowLastUsed, 8, RightTotalWeight)
	RowLastUsed += 1

	sheet1.write(RowLastUsed, 0, 'Substitute Paddlers: ')
	RowLastUsed += 1
	sheet1.write(RowLastUsed, 0, 'Name: ')
	sheet1.write(RowLastUsed, 1, 'Weight: ')
	sheet1.write(RowLastUsed, 2, 'Gender: ')
	sheet1.write(RowLastUsed, 3, 'Side: ')
	sheet1.write(RowLastUsed, 4, 'Trial Time: ')
	sheet1.write(RowLastUsed, 5, 'Notes: ')
	RowLastUsed += 1
	for m in range(0, len(Boats[i].Subs)):
		sheet1.write(m + RowLastUsed, 0, Boats[i].Subs[m].Name)
		sheet1.write(m + RowLastUsed, 1, Boats[i].Subs[m].Weight)
		sheet1.write(m + RowLastUsed, 2, Boats[i].Subs[m].Gender)
		sheet1.write(m + RowLastUsed, 3, Boats[i].Subs[m].Side)
		sheet1.write(m + RowLastUsed, 4, Boats[i].Subs[m].TimeTrial)
		sheet1.write(m + RowLastUsed, 5, Boats[i].Subs[m].Notes)
	RowLastUsed += len(Boats[i].Subs)
	RowLastUsed += 1

book.save("Results.xls")

# print('Left Paddlers:')
# for i in range (0, len(BoatLeft)):
# 	BoatLeft[i].printPaddler()
# print('Right Paddlers:')
# for i in range (0, len(BoatRight)):
# 	BoatRight[i].printPaddler()
# print('Both: ')
# for i in range (0, len(BoatBoth)):
# 	BoatBoth[i].printPaddler()

# print(sheet.cell(row=j, column=i).value)

# writer = pd.ExcelWriter('List.xlsx', engine='xlsxwriter')

# yourData.to_excel(writer, 'Sheet1')

# writer.save()